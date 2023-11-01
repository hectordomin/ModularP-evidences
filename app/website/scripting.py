from planrcomm.asgi import *
import openpyxl
from core.erp.models import Subject
from core.erp.models import Professor

def populate_subjects():
    data = "/home/devusr/project/ModularP-evidences/resources/data.xlsx"
    wb = openpyxl.load_workbook(data)
    sheet = wb.active

    for value in sheet.iter_rows(min_col=2, max_col=2, values_only=True):
        if value[0] != None:
            subject = Subject(name= value[0],)
            #print(value[0])
            subject.save()

    print("\n Tabla de materias llenada éxitosamente")

def populate_professors():
    data = "/home/devusr/project/ModularP-evidences/resources/data.xlsx"
    wb = openpyxl.load_workbook(data)
    sheet = wb.active

    for value in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
        professor = Professor(name= value[0],)
        #print(value[0])
        professor.save()

    print("\n Tabla de profesores llenada éxitosamente")

populate_subjects()
populate_professors()