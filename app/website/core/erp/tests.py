from django.test import TestCase
import openpyxl
from .models import Subject
from .models import Professor

# Create your tests here.

class testSubject(TestCase):
    
    def populate_subjects(self):
        data = "/home/devusr/project/ModularP-evidences/resources/data.xlsx"
        wb = openpyxl.load_workbook(data)
        sheet = wb.active

        for value in sheet.iter_rows(min_col=2, max_col=2, values_only=True):
            if value[0] != None:
                subject = Subject(name= value[0],)
                print(value[0])
                subject.save()

        print("\n Tabla de materias llenada éxitosamente")

    def add_subject(self):
        subject = Subject(
            name='Juan',
        )
        subject.save()
        print("Materia creada exitosamente")

class testProfessor(TestCase):
    
    def populate_professors(self):
        data = "/home/devusr/project/ModularP-evidences/resources/data.xlsx"
        wb = openpyxl.load_workbook(data)
        sheet = wb.active

        for value in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
            professor = Professor(name= value[0],)
            print(value[0])
            professor.save()

        print("\n Tabla de profesores llenada éxitosamente")